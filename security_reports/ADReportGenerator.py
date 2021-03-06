import openpyxl, os, sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from win32com import client
# The following can be uncommented if files need to be converted from CSV to XLSX
# import pyexcel
'''
# Save the csv as xlsx
sheet = pyexcel.get_sheet(file_name="test.csv", delimiter=",")
sheet.save_as("test.xlsx")
'''
# List of input files with the following order: Input Filename, Output Filename, Output Title, Column of Name, Column of Boolean, Boolean to Flag
files = [["ActiveUsers", "AD_Active_Users.xlsx", "AD Active Users", "A", "D", "TRUE", "AD_Active_Users.pdf","A"], ["Admin", "AD_Admin_Users.xlsx", "AD Admin Users", "B", "D", "FALSE", "AD_Admin_Users.pdf","B"], ["Computers", "AD_Active_Computers.xlsx", "AD Active Computers", "A", "C", "TRUE", "AD_Active_Computers.pdf","A"], ["SslVpn", "AD_SSLVPN_Users.xlsx", "AD SSL/VPN Users", "B", "D", "FALSE", "AD_SSLVPN_Users.pdf","B"]]

# Find the input filenames using startswith, then remove items from 'files' if the file does not exist.
def add_fnames(files):
    local_dir = os.listdir()                                            # Create a list of all the files in the directory
    i = 0
    no_match = []
    for item in files:                                                  # Grab a list item from 'files'
        match = 0
        for filename in local_dir:                                      # Loop thru all of the filenames in the directory
            if filename.startswith(item[0]):                            # Check for a match to the first sublist value in the list item
                item[0] = filename
                match += 1
        if match == 0:
            no_match.append(i)                                          # If there is no matching file, remove that list item
        i += 1
    no_match.reverse()                                                  # Reverse the list of indexes of the items to remove
    for item in no_match:                                               # Delete all list items which have no corresponding file in the directory
        del files[item]

# Highlight items, add status, and comment on status
def highlight(column_count, row_count, file):
    ws[(chr(65 + column_count)) + "1"] = "Status"                       # Add a Status column

    # Highlight the first row
    for i in range(0,column_count + 1):
        cell = chr(65 + i) + "1"
        ws[cell].fill = PatternFill("solid", start_color="FFAF6E")

    # Highlight the item being looked at
    for i in range(2,row_count + 1):
        cell = file[7] + str(i)
        ws[cell].fill = PatternFill("solid", start_color="E1FFE0")

    # Highlight all instances of the correct Boolen value (as found in the 'files' list) and query status
    for i in range(1,row_count + 1):
        cell = file[4] + str(i)
        true_false = str(ws[cell].value)
        if true_false.lower() == (file[5]).lower():
            os.system('cls')
            ws[cell].fill = PatternFill("solid", start_color="FFFF75")
            item = ws[file[3] + str(i)].value
            while True:
                print(file[2])
                status = input(f"Enter the status for {item}.\n1) Keep\n2) Remove\n3) Review\n>")
                if status == "1" or status == "2" or status == "3":
                    if status == "1":
                        ws[(chr(65 + column_count)) + str(i)] = "Keep"
                        break
                    elif status == "2":
                        ws[(chr(65 + column_count)) + str(i)] = "Remove"
                        break
                    elif status == "3":
                        ws[(chr(65 + column_count)) + str(i)] = "Review"
                        break
                else:
                    print("PLEASE ENTER A VALID OPTION\n-----------------------------------------------------------------------")


# Center all items
def center(column_count, row_count):
    widths = []
    for col in range(0, column_count + 1):
        this_col = chr(65 + col)
        length = 0
        for row in range(1, row_count + 1):
            this_cell = this_col + str(row)
            ws[this_cell].alignment = Alignment(horizontal="center")
            if ws[this_cell].value != None and len(str(ws[this_cell].value)) > length:
                length = len(str(ws[this_cell].value))
        widths.append(length)
    return widths

# Change the width of the columns to match the longest item in the list    
def col_space(widths):
    i = 0
    for item in widths:
        ws.column_dimensions[chr(65 + i)].width = item + 1
        i += 1

# Add rows at the top, merge those rows and add title
def title(file, column_count):
    ws.insert_rows(1,2)
    last_col = chr(65 + column_count)
    end_block_a = last_col + "2"
    ws.merge_cells(f"A1:{end_block_a}")
    ws["A1"] = file[2]
    ws["A1"].font = Font(name="Calibri", size="18")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", start_color="E0F4FF")

# Convert to PDF
def pdf_convert(files):
    # Get local directory path
    environment = os.getcwd() + "\\"
    # Get the output filenames from the 'files' list
    for item in files:
        raw_file = environment + item[0]                                            # Concatenate path to starting file
        input_path = environment + item[1]                                          # Concatenate path to working file
        output_path = environment + item[6]                                         # Concatenate path to output file
        excel = client.Dispatch("Excel.Application")                                # Open Microsoft Excel
        sheets = excel.Workbooks.Open(input_path)                                   # Read Excel File
        work_sheets = sheets.Worksheets[0]                                          # Read sheet 1
        work_sheets.ExportAsFixedFormat(0, output_path)                             # Convert into PDF File
        sheets.Close(True)                                                          # Close the workbook
        os.remove(raw_file)
        os.remove(input_path)

### MAIN ###
# Change to the actual filenames of the input files
add_fnames(files)
if files == []:                                                                     # End the program if there are no matching files
    sys.exit()

for file in files:
    wb = load_workbook(file[0])                                                     # Load the workbook
    ws = wb.active
    column_count = ws.max_column                                                    # Calculate the number of rows and columns
    row_count = ws.max_row
    highlight(column_count, row_count, file)
    widths = center(column_count, row_count)                                        # Center all fields and return the max width for each column
    col_space(widths)
    title(file, column_count)
    ws.sheet_properties.pageSetUpPr.fitToPage = True                                # Fit to page 
    ws.page_setup.fitToHeight = False
    wb.save(file[1])
pdf_convert(files)

# pyinstaller.exe --onefile -i"icon_path" py_file_path
# After running pyinstaller, change Analysis->pathex in the spec file to the location of the site-packages ("C:\Python310\Lib\site-packages" in my case)
# Rerun "pyinstaller.exe file_name.spec"