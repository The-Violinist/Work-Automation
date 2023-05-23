import pyexcel
import openpyxl, os, sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from win32com import client

# List of input files with the following order: CSV Filename, XLSX Filename, Formatted XLSX Filename, Output Title, PDF Filename
files = [
    [
        "Authentication_Allowed",
        "Auth Allowed Temp.xlsx",
        "SSLVPN Authentication Allowed.xlsx",
        "SSLVPN Authentication Allowed",
        "SSLVPN Authentication Allowed.pdf",
        "Authentication_Allowed.csv",
    ],
    [
        "Authentication_Denied",
        "Auth Denied Temp.xlsx",
        "SSLVPN Authentication Denied.xlsx",
        "SSLVPN Authentication Denied",
        "SSLVPN Authentication Denied.pdf",
        "Authentication_Denied.csv",
    ],
]

# Find the input filenames using startswith, then remove items from 'files' if the file does not exist.
def add_fnames(files):
    local_dir = os.listdir()  # Create a list of all the files in the directory
    i = 0
    no_match = []
    for item in files:  # Grab a list item from 'files'
        match = 0
        for filename in local_dir:  # Loop thru all of the filenames in the directory
            if filename.__contains__(item[0]) and filename.__contains__(
                ".csv"
            ):  # Check for a match to the first sublist value in the list item
                os.rename(filename, item[5])
                match += 1
        if match == 0:
            no_match.append(i)  # If there is no matching file, remove that list item
        i += 1
    no_match.reverse()  # Reverse the list of indexes of the items to remove
    for (
        item
    ) in (
        no_match
    ):  # Delete all list items which have no corresponding file in the directory
        del files[item]


def convertCSV():
    i = 0
    for item in files:
        try:
            sheet = pyexcel.get_sheet(file_name=item[5], delimiter=",")
            sheet.save_as(item[1])
            os.remove(item[5])
        except:
            pass
        i += 1


# Highlight first row
def highlight(column_count):
    # Highlight the first row
    for i in range(0, column_count + 1):
        cell = chr(65 + i) + "1"
        ws[cell].fill = PatternFill("solid", start_color="FFAF6E")


# Center all items
def center(column_count, row_count):
    widths = []
    for col in range(0, column_count + 1):
        this_col = chr(65 + col)
        length = 0
        for row in range(1, row_count + 1):
            this_cell = this_col + str(row)
            ws[this_cell].alignment = Alignment(horizontal="center")
            if ws[this_cell].value != None and len(str(ws[this_cell].value)) > length:
                length = len(str(ws[this_cell].value))
        widths.append(length)
    return widths


# Change the width of the columns to match the longest item in the list
def col_space(widths):
    i = 0
    for item in widths:
        ws.column_dimensions[chr(65 + i)].width = item + 3
        i += 1


# Add rows at the top, merge those rows and add title
def title(file, column_count):
    ws.insert_rows(1, 2)
    last_col = chr(65 + column_count)
    end_block_a = last_col + "2"
    ws.merge_cells(f"A1:{end_block_a}")
    ws["A1"] = file[3]
    ws["A1"].font = Font(name="Calibri", size="18")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", start_color="E0F4FF")


def wbsorted(filename):
    environment = os.getcwd() + "\\"
    excel = client.Dispatch("Excel.Application")
    workb = excel.Workbooks.Open(environment + filename)
    works = workb.Worksheets[0]
    if filename.__contains__("Allowed"):
        works.Range("F:F").Delete()
    used = works.UsedRange
    nrows = used.Row + used.Rows.Count - 1
    ncols = used.Column + used.Columns.Count - 1
    range_end = chr(65 + ncols) + str(nrows)
    works.Range(f"A2:{range_end}").Sort(Key1=works.Range("A1"), Order1=1, Orientation=1)

    workb.Save()
    excel.Application.Quit()


# Convert to PDF
def pdf_convert(files):
    # Get local directory path
    environment = os.getcwd() + "\\"
    # Get the output filenames from the 'files' list
    for item in files:
        input_path = environment + item[2]  # Concatenate path to working file
        output_path = environment + item[4]  # Concatenate path to output file
        excel = client.Dispatch("Excel.Application")  # Open Microsoft Excel
        sheets = excel.Workbooks.Open(input_path)  # Read Excel File
        work_sheets = sheets.Worksheets[0]  # Read sheet 1
        work_sheets.PageSetup.Orientation = 2
        work_sheets.ExportAsFixedFormat(0, output_path)  # Convert into PDF File
        sheets.Close(True)  # Close the workbook
        os.remove(input_path)


### MAIN ###

add_fnames(files)
convertCSV()

for file in files:
    wbsorted(file[1])
    wb = load_workbook(file[1])  # Load the workbook
    ws = wb.active
    column_count = ws.max_column  # Calculate the number of rows and columns
    row_count = ws.max_row
    highlight(column_count)
    widths = center(
        column_count, row_count
    )  # Center all fields and return the max width for each column
    col_space(widths)
    title(file, column_count)
    ws.sheet_properties.pageSetUpPr.fitToPage = True  # Fit to page
    ws.page_setup.fitToHeight = False
    wb.save(file[2])
    os.remove(file[1])
pdf_convert(files)
